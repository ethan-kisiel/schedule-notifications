import openpyxl
from datetime import datetime
from twilio.rest import Client

class Schedule:
    '''
    '''
    def __init__(self, schedule: str, limits):
        self.schedule = schedule
        self.schedule_data = self.get_data(*limits)
        self.days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        
    def get_data(self, title_row: int, title_col, min_col: int, max_col: int, min_row: int, max_row: int):
        '''
        '''
        try:
            wb = openpyxl.load_workbook(self.schedule, data_only=True)
            ws = wb.active
            sched_dict = {}
            # set up dictionary with day values
            for col in range(min_col, max_col):
                curr_row_title = ws.cell(row=title_row, column=col).value
                if curr_row_title != None:
                    sched_dict[curr_row_title] = dict()
                for row in range(min_row, max_row):
                    curr_cell = ws.cell(row=row, column=col).value
                    curr_col_title = ws.cell(row=row, column=title_col).value
                    if curr_cell != None:
                        sched_dict[curr_row_title][self.convert_time(curr_col_title)] = curr_cell
            return sched_dict
        except FileNotFoundError:
            print('ERROR: file not found.')

    def convert_time(self, time: datetime.time):
        '''
        Takes datetime.time
        returns string in 12 hour format
        '''
        time = str(time).split(':')
        if len(time) > 2:
            time.pop(2)
        if int(time[0]) > 12:
            time[0] = str(int(time[0]) - 12)
            return(f'{time[0]}:{time[1]}:PM')
        else:
            return(f'{time[0]}:{time[1]}:AM')

    def compare_time(self, time: str, buffer: int):
        '''
        Takes time in the format of HH:MM:AM/PM
        returns true if time is within buffer(minutes) of current time
        '''
        current_time = datetime.now()
        current_time = self.convert_time(current_time.strftime('%H:%M')).split(':')
        time = time.split(':')
        if current_time[2] != time[2]:
            return False
        elif int(time[0]) - int(current_time[0]) == 1 and int(current_time[1]) >= int(time[1] + (60 - buffer)):
            return False
        elif int(current_time[1]) <= int(time[1]) and int(current_time[1]) >= int(time[1]) - buffer:
            return True
        else:
            return False

    def get_day(self):
        today = datetime.today().weekday()
        return self.schedule_data[self.days[today]]
    
    def send_notification(self, twilio_sid, twilio_auth, to_number, from_number, msg):
        '''
        Takes twilio account sid, twilio authentication token, target number, twilio number, message body
        '''
        try:
            client = Client(twilio_sid, twilio_auth)
            message = client.messages.create(
                to=to_number,
                from_=from_number,
                body=msg
            )

        except Exception as e:
            print(e)
